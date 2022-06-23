from mimetypes import init
from openpyxl import Workbook, load_workbook
import re
import json
import copy

json_data = {"courses":[]} # The data we will be storing!

cols_name = ["A","B","C","D","E","F","G","H","I","J","K"]

json_data_entry_keys = ["dept","course_name","slot_name","credits","course_type","instructor","instructor_email","dh","tut","practical"]

n_courses_added = 0 # To keep track of the number of courses extracted.

n_row = 2

def is_row_empty(df,row_index):
    # On giving the current sheet, the names of the columns, and the row to be checked. 
    # Returns True if it is empty & False if it is non-empty.
    global cols_name
    for x in cols_name:
        if not df[f"{x}{row_index}"].value==None:
            return False
    return True
    

def get_row_values(df,row_index):
    # Returns the values of the entries in each column corresponding to the as a list 
    global cols_name
    out = []
    for x in cols_name:
        out.append(df[f"{x}{row_index}"].value)
    return out

def extract_fields(df):
    global n_courses_added
    global n_row
    tmp_dict = {}
    tmp_dict["n_course"] = n_courses_added+1
    tmp_dict["dept"]=str(df[f"B{n_row}"].value).strip()
    tmp_dict["course_name"]=str(df[f"C{n_row}"].value).strip()
    tmp_dict["slot_name"]=str(df[f"D{n_row}"].value).strip()
    tmp_dict["credits"]=str(df[f"E{n_row}"].value).strip()
    tmp_dict["course_type"]=str(df[f"F{n_row}"].value).strip()
    tmp_dict["instructor"]=str(df[f"G{n_row}"].value).strip()
    tmp_dict["instructor_email"]=str(df[f"H{n_row}"].value).strip()
    tmp_dict["dh"]=str(df[f"I{n_row}"].value).strip()
    tmp_dict["tut"]=str(df[f"J{n_row}"].value).strip()
    tmp_dict["practical"]=str(df[f"K{n_row}"].value).strip()
    return tmp_dict

def extract_and_append_current_row(df):
    global json_data
    global n_courses_added
    global n_row

    tmp_dict = extract_fields(df)
    json_data["courses"].append(tmp_dict)
    n_courses_added+=1
    n_row+=1

def show_differences_in_row(row_val_list, init_val, final_val):
    for ij in range(1, len(row_val_list) ):
        mykey = json_data_entry_keys[ij-1]
        if final_val[mykey] != init_val[mykey]: # If this field has changed => Display.
            print(f"Value for {mykey} changed. | {init_val[mykey]} ==> {final_val[mykey]}")
    print("#################################################")


def append_missing_data(df):
    global json_data
    global json_data_entry_keys
    global n_courses_added
    global n_row
    row_val_list = get_row_values(df,n_row)
    print(f"#################################################\nFound an non-empty row n_row = {n_row} with missing S.No. entry : {row_val_list}\n")

    # print(f"Value of the prev json_data entry : {json_data['courses'][n_courses_added-1]}",end="\n\n")

    initial_value = (json_data['courses'][n_courses_added-1]).copy() # To check the diff 

    prev_index = n_courses_added-1
    for ij in range(1, len(row_val_list) ):
        str_to_append = " "+str(row_val_list[ij]).strip()
        if str_to_append.strip() == 'None' or str_to_append==None:
            continue
        json_data["courses"][prev_index][json_data_entry_keys[ij-1]] += str_to_append

    show_differences_in_row(row_val_list, initial_value, json_data['courses'][n_courses_added-1])
    # print(f"Value of the prev json_data entry After MOD : {json_data['courses'][n_courses_added-1]}")
    # print("\n#####################")
    n_row+=1

def extract_course_code_and_append():
    global json_data
    listOfCourses = json_data['courses']
    for i in range(len(listOfCourses)):
        courseName = listOfCourses[i]['course_name']
        try:
            m = re.match(r"[a-zA-Z0-9\-&.,:\s]*\((?P<course_code>[A-Z0-9\s]+)\)", courseName)
            courseCode = str(m.groupdict()['course_code'])
            courseCode = courseCode.replace(" ","")
        except:
            print("ERROR : Couldn't parse course code for course : ",courseName)
            courseCode = "PARSING_ERROR"

        listOfCourses[i]['course_code'] = courseCode
    json_data["courses"] = listOfCourses


def extract_data_from_current_sheet(wb, sheet_name):
    global n_row
    global cols_name
    global n_courses_added
    print(f"\n********* Working on sheet {sheet_name} *********\n")
    df = wb[sheet_name]

    n_row = 2
    
    while(1):
        
        if is_row_empty(df,n_row): #? Check if the entry is completely empty i.e THE CURRENT SHEET HAS BEEN COVERED.
            print(f"Moving to the next sheet, empty row found at {n_row}")
            break
        
        # Check if there is some data that has been pushed to a next row from the previous one.
        # And we will log these incidents, as these might be a source of errors.
        if df[f"A{n_row}"].value == None:
            append_missing_data(df)
            continue
        
        # print(f"Entry Number as mentioned in the sheet : {df[f'A{n_row}'].value} | {type(df[f'A{n_row}'].value)}")
        if not type(df[f'A{n_row}'].value) == int:
            n_row+=1
            continue

        # If the entry is normal, just load the data.
        extract_and_append_current_row(df)

def save_json_data(filename):
    global json_data
    with open((filename.split('/')[-1])[:-5]+".json", 'w') as fp:
        json.dump(json_data, fp)

def extract_records(excel_filename):
    wb = load_workbook(filename=excel_filename) # Load the excel file

    for sheet_name in wb.sheetnames:
        extract_data_from_current_sheet(wb, sheet_name)
    print(f"Extracted {n_courses_added} courses, VERIFY the last course's serial number with {n_courses_added}")
    print(f"STARTING THE COURSE_CODE EXTRACTION...")
    extract_course_code_and_append()
    save_json_data(excel_filename)
    print(f"Finished extracting the records! Bye!")

if __name__ == "__main__":
    # Get the arguments passed along with the command.
    

    extract_records("Course_Schedule_2022-23-1.xlsx")
